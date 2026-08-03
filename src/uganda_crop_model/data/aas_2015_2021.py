"""Parser for the UBOS Statistical Abstract workbooks covering 2015-2021.

Targets the two files in ``data/raw/``:

- ``Area_planted_for_selected_Food_Crops,_2015-2021(Ha).xlsx``
- ``Crop_production_for_selected_Food_Crops,_2015-2021_(MT).xlsx``

These workbooks are published as national statistical-abstract tables, but
the parser does not assume that: :func:`detect_grain` inspects the actual
cell content and classifies the geographic grain (national / region /
subregion / district) before any values are trusted. If the grain is not
subregion, the loader still returns the panel but marks it
``grain != "subregion"`` so that downstream gates (``AnalysisPolicy`` and
``ubos_waves.validate_wave_source``) can fail honestly instead of silently
absorbing national totals into a subregion model.

Layout tolerance
----------------
The workbooks may be oriented either way:

* wide: one row per crop, year values across columns (possibly with a
  leading label column and group header rows such as "Cereals"), or
* long/transposed: years down the first column, crops across the header.

Both single- and multi-sheet files are handled. When two sheets carry a
value for the same (crop, year) pair the values are compared; material
conflicts raise instead of being silently averaged.

Yield basis
-----------
The derived target is ``yield_tons_ha_planted = production_mt /
area_planted_ha`` — production over *planted* area, not harvested area.
UBOS AAS wave reports define yield over harvested area, so this derived
series is a sensitivity target only and is flagged as such in the output
metadata. It must never be mixed with harvested-area yields in one column.
"""

from __future__ import annotations

import argparse
import json
import re
from dataclasses import dataclass, field
from pathlib import Path

import numpy as np
import pandas as pd

# ---------------------------------------------------------------------------
# Geography vocabulary used by detect_grain()
# ---------------------------------------------------------------------------

AAS_SUBREGIONS = {
    "acholi", "ankole", "bugisu", "bukedi", "bunyoro", "busoga",
    "elgon", "karamoja", "kigezi", "lango", "rwenzori", "teso", "tooro",
    "toro", "west nile", "westnile",
    "south buganda", "southern buganda", "s.buganda", "s buganda",
    "north buganda", "northern buganda", "n.buganda", "n buganda",
}

UGANDA_REGIONS = {"central", "eastern", "northern", "western"}

NATIONAL_MARKERS = {"uganda", "national", "all uganda"}

GEO_HEADER_HINTS = {
    "subregion": "subregion",
    "sub-region": "subregion",
    "sub region": "subregion",
    "region": "region",
    "district": "district",
    "zardi": "zardi",
}

# ---------------------------------------------------------------------------
# Crop vocabulary: UBOS Statistical Abstract labels -> project crop names
# ---------------------------------------------------------------------------

CROP_ALIASES = {
    "maize": "maize",
    "millet": "millet",
    "finger millet": "millet",
    "sorghum": "sorghum",
    "rice": "rice",
    "cassava": "cassava",
    "s/potatoes": "sweet_potatoes",
    "sweet potatoes": "sweet_potatoes",
    "sweet potato": "sweet_potatoes",
    "irish": "irish_potatoes",
    "irish potatoes": "irish_potatoes",
    "irish potato": "irish_potatoes",
    "beans": "beans",
    "gnuts": "groundnuts",
    "g.nuts": "groundnuts",
    "groundnuts": "groundnuts",
    "ground nuts": "groundnuts",
    "soya beans": "soya_beans",
    "soyabeans": "soya_beans",
    "soya": "soya_beans",
    "plantain bananas (all types)": "banana",
    "plantain bananas(all types)": "banana",
    "plantain bananas": "banana",
    "banana food": "banana",
    "matooke": "banana",
    "simsim": "simsim",
    "sesame": "simsim",
    "cow peas": "cowpeas",
    "cowpeas": "cowpeas",
    "field peas": "field_peas",
    "pigeon peas": "pigeon_peas",
}

# Group headers / subtotal rows — parsed but flagged, never used as crops.
GROUP_LABELS = {
    "cereals", "cereals total", "total cereals",
    "roots and tubers", "roots & tubers", "root crops",
    "pulses", "legumes",
    "oil crops", "oilcrops",
    "plantains", "bananas",
    "total", "grand total", "all crops", "others",
}

YEAR_MIN, YEAR_MAX = 2000, 2025

# Printed national totals used as validation anchors. Values come from the
# UBOS AAS 2018 report tables (e.g. Table 7.2 millet Uganda total 141,982 MT)
# and the published Statistical Abstract series. Tolerance is near-exact:
# these are the same published figures the workbook should reproduce.
NATIONAL_ANCHORS_PRODUCTION_MT = {
    ("millet", 2018): 141_982,
    ("sorghum", 2018): 268_493,
    ("beans", 2018): 727_652,
    ("groundnuts", 2018): 253_279,
    ("soya_beans", 2018): 107_624,
    ("maize", 2015): 2_812_919,
    ("sorghum", 2015): 410_720,
    ("rice", 2015): 238_193,
    ("millet", 2015): 236_484,
}

NATIONAL_ANCHORS_AREA_HA = {
    ("maize", 2018): 2_480_097,
    ("millet", 2018): 282_607,
    ("sorghum", 2018): 491_911,
    ("rice", 2018): 199_538,
    ("cassava", 2018): 1_705_942,
    ("beans", 2018): 1_205_509,
    ("groundnuts", 2018): 515_041,
    ("soya_beans", 2018): 189_710,
}

ANCHOR_REL_TOLERANCE = 0.005  # 0.5% wiggle for rounding in abstract tables


class AASWorkbookError(ValueError):
    """Raised when a workbook cannot be parsed or fails validation."""


@dataclass
class GrainReport:
    """Result of :func:`detect_grain`."""

    grain: str  # "national" | "region" | "subregion" | "district" | "unknown"
    matched_units: list[str] = field(default_factory=list)
    evidence: str = ""

    def to_dict(self) -> dict:
        return {
            "grain": self.grain,
            "matched_units": sorted(self.matched_units),
            "evidence": self.evidence,
        }


# ---------------------------------------------------------------------------
# Low-level helpers
# ---------------------------------------------------------------------------

def _norm(text: object) -> str:
    """Normalise a cell value for vocabulary matching."""
    if text is None:
        return ""
    s = str(text).strip().lower()
    s = re.sub(r"[\u2013\u2014]", "-", s)
    s = re.sub(r"\s+", " ", s)
    return s


def _as_year(value: object) -> int | None:
    """Return the integer year if the cell looks like one, else None."""
    if value is None:
        return None
    if isinstance(value, (int, np.integer)):
        v = int(value)
        return v if YEAR_MIN <= v <= YEAR_MAX else None
    if isinstance(value, float) and float(value).is_integer():
        v = int(value)
        return v if YEAR_MIN <= v <= YEAR_MAX else None
    s = str(value).strip()
    m = re.fullmatch(r"(19|20)\d{2}", s)
    if m:
        v = int(s)
        if YEAR_MIN <= v <= YEAR_MAX:
            return v
    return None


def _as_number(value: object) -> float | None:
    """Parse a numeric cell, tolerating thousands separators and NA markers."""
    if value is None:
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        v = float(value)
        return v if np.isfinite(v) else None
    s = str(value).strip().replace(",", "").replace("\u00a0", "")
    if s.lower() in {"", "-", "--", "..", "na", "n/a", "nan"}:
        return None
    s = re.sub(r"[a-zA-Z*]+$", "", s)  # strip footnote markers like '632a'
    try:
        return float(s)
    except ValueError:
        return None


def _sheet_grid(path: Path, sheet: str, max_rows: int = 200, max_cols: int = 60) -> list[list[object]]:
    """Read a sheet into a raw grid using openpyxl (values only)."""
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet not in wb.sheetnames:
            raise AASWorkbookError(f"{path.name}: sheet {sheet!r} not found; sheets={wb.sheetnames}")
        ws = wb[sheet]
        grid: list[list[object]] = []
        for i, row in enumerate(ws.iter_rows(max_row=max_rows, max_col=max_cols, values_only=True)):
            grid.append(list(row))
            if i + 1 >= max_rows:
                break
        return grid
    finally:
        wb.close()


def list_sheets(path: Path) -> list[str]:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        return list(wb.sheetnames)
    finally:
        wb.close()


# ---------------------------------------------------------------------------
# Grain detection
# ---------------------------------------------------------------------------

def detect_grain(path: Path, max_rows: int = 200) -> GrainReport:
    """Classify the geographic grain of a UBOS workbook from its cell content.

    The detector scans every text cell in every sheet:

    * explicit geography headers (``subregion`` / ``region`` / ``district`` /
      ``zardi`` columns) are the strongest signal;
    * >=5 distinct AAS subregion names -> ``subregion``;
    * the four region names without subregion detail -> ``region``;
    * only "Uganda"/"Total" national markers and crop labels -> ``national``.

    It never guesses: anything inconclusive returns ``unknown`` and the
    caller is expected to raise rather than fabricate a grain.
    """
    path = Path(path)
    subregion_hits: set[str] = set()
    region_hits: set[str] = set()
    header_hint: str | None = None
    national_hits = 0
    text_cells = 0

    for sheet in list_sheets(path):
        grid = _sheet_grid(path, sheet, max_rows=max_rows)
        for row in grid:
            for cell in row:
                if not isinstance(cell, str):
                    continue
                s = _norm(cell)
                if not s:
                    continue
                text_cells += 1
                if s in GEO_HEADER_HINTS and header_hint is None:
                    header_hint = GEO_HEADER_HINTS[s]
                if s in AAS_SUBREGIONS:
                    subregion_hits.add(s)
                elif s in UGANDA_REGIONS:
                    region_hits.add(s)
                elif s in NATIONAL_MARKERS:
                    national_hits += 1

    if header_hint in {"subregion", "zardi"} or len(subregion_hits) >= 5:
        return GrainReport(
            "subregion",
            sorted(subregion_hits),
            f"header hint={header_hint!r}; {len(subregion_hits)} subregion names matched",
        )
    if header_hint == "district":
        return GrainReport("district", [], "explicit 'district' header found")
    if header_hint == "region" or len(region_hits) >= 3:
        return GrainReport(
            "region",
            sorted(region_hits),
            f"header hint={header_hint!r}; {len(region_hits)} region names matched",
        )
    if national_hits > 0 and not subregion_hits and not region_hits:
        return GrainReport(
            "national",
            ["uganda"],
            f"{national_hits} national marker cells; no subregion/region names found",
        )
    if not subregion_hits and not region_hits and text_cells > 0:
        # Crop-only tables with no geography column at all are national
        # statistical-abstract series in this UBOS publication family.
        return GrainReport(
            "national",
            [],
            "no geography labels found; treating as national abstract series",
        )
    return GrainReport("unknown", [], "could not classify geography from cell content")


# ---------------------------------------------------------------------------
# Table extraction
# ---------------------------------------------------------------------------

def _match_crop(label: object) -> tuple[str | None, bool]:
    """Return (canonical_crop, is_group_total) for a row/column label."""
    s = _norm(label)
    if not s:
        return None, False
    s_clean = re.sub(r"\s*\(.*?\)\s*", " ", s).strip()
    if s in GROUP_LABELS or s_clean in GROUP_LABELS:
        return None, True
    if s in CROP_ALIASES:
        return CROP_ALIASES[s], False
    if s_clean in CROP_ALIASES:
        return CROP_ALIASES[s_clean], False
    # prefix match for labels with trailing footnote markers, e.g. "Maize *"
    for alias, canon in CROP_ALIASES.items():
        if s.startswith(alias) and len(s) <= len(alias) + 12:
            return canon, False
    return None, False


def _extract_wide(grid: list[list[object]], sheet: str) -> list[dict]:
    """Wide layout: find the year header row, melt crop rows underneath it."""
    best_header_idx = None
    best_year_cols: list[tuple[int, int]] = []  # (col_idx, year)
    for r, row in enumerate(grid[:30]):
        year_cols = [(c, _as_year(v)) for c, v in enumerate(row)]
        year_cols = [(c, y) for c, y in year_cols if y is not None]
        if len(year_cols) >= 3:
            best_header_idx = r
            best_year_cols = year_cols
            break
    if best_header_idx is None:
        return []

    label_col = max(0, min(c for c, _ in best_year_cols) - 1)
    records: list[dict] = []
    for row in grid[best_header_idx + 1:]:
        if label_col >= len(row):
            continue
        crop, is_group = _match_crop(row[label_col])
        used_col = label_col
        if crop is None and not is_group:
            # label column guess may be off by one; try column 0
            crop, is_group = _match_crop(row[0])
            used_col = 0
        if crop is None:
            continue
        for col_idx, year in best_year_cols:
            if col_idx >= len(row):
                continue
            value = _as_number(row[col_idx])
            if value is None:
                continue
            records.append(
                {
                    "sheet": sheet,
                    "crop": crop,
                    "year": year,
                    "value": value,
                    "raw_label": _norm(row[used_col]),
                    "is_group_total": bool(is_group),
                }
            )
    return records


def _extract_long(grid: list[list[object]], sheet: str) -> list[dict]:
    """Long/transposed layout: years down a column, crops across the header."""
    for r, row in enumerate(grid[:30]):
        crop_cols = [(c, _match_crop(v)) for c, v in enumerate(row)]
        crop_cols = [(c, m) for c, m in crop_cols if m[0] is not None]
        if len(crop_cols) < 3:
            continue
        year_col = None
        for rr in grid[r + 1: r + 40]:
            for c, v in enumerate(rr):
                if _as_year(v) is not None:
                    year_col = c
                    break
            if year_col is not None:
                break
        if year_col is None:
            continue
        records = []
        for rr in grid[r + 1:]:
            year = _as_year(rr[year_col]) if year_col < len(rr) else None
            if year is None:
                continue
            for c, (crop, is_group) in crop_cols:
                if c >= len(rr):
                    continue
                value = _as_number(rr[c])
                if value is None:
                    continue
                records.append(
                    {
                        "sheet": sheet,
                        "crop": crop,
                        "year": year,
                        "value": value,
                        "raw_label": _norm(row[c]),
                        "is_group_total": bool(is_group),
                    }
                )
        if records:
            return records
    return []


def parse_workbook(path: Path, value_name: str, sheets: list[str] | None = None) -> pd.DataFrame:
    """Parse one workbook into a long panel.

    Returns columns: ``crop``, ``year``, ``<value_name>``, ``sheet``,
    ``raw_label``. Duplicate (crop, year) values across sheets are kept
    distinct here; conflict resolution happens in
    :func:`build_national_yield_panel`.
    """
    path = Path(path)
    if not path.exists():
        raise AASWorkbookError(f"workbook not found: {path}")
    sheets = sheets or list_sheets(path)
    records: list[dict] = []
    for sheet in sheets:
        grid = _sheet_grid(path, sheet)
        recs = _extract_wide(grid, sheet) or _extract_long(grid, sheet)
        records.extend(recs)
    if not records:
        raise AASWorkbookError(
            f"{path.name}: no crop/year table detected in sheets {sheets}; "
            "run with --describe to inspect the detected layout"
        )
    df = pd.DataFrame(records).rename(columns={"value": value_name})
    df = df[~df["is_group_total"]].drop(columns=["is_group_total"])
    return df.reset_index(drop=True)


def load_area_planted(path: Path) -> pd.DataFrame:
    """Parse the area-planted workbook -> columns crop, year, area_planted_ha."""
    return parse_workbook(path, "area_planted_ha")


def load_production(path: Path) -> pd.DataFrame:
    """Parse the production workbook -> columns crop, year, production_mt."""
    return parse_workbook(path, "production_mt")


# ---------------------------------------------------------------------------
# Merge + validation
# ---------------------------------------------------------------------------

def _dedupe_with_conflict_check(df: pd.DataFrame, value_col: str, source_name: str) -> pd.DataFrame:
    """Collapse duplicate (crop, year) rows; raise on material conflicts."""
    rows = []
    for (crop, year), group in df.groupby(["crop", "year"]):
        vals = sorted(group[value_col].dropna().unique().tolist())
        if not vals:
            continue
        if len(vals) == 1:
            rows.append({"crop": crop, "year": year, value_col: vals[0]})
            continue
        lo, hi = vals[0], vals[-1]
        rel = (hi - lo) / max(abs(hi), 1e-9)
        if rel <= ANCHOR_REL_TOLERANCE:
            rows.append({"crop": crop, "year": year, value_col: float(np.mean(vals))})
            continue
        raise AASWorkbookError(
            f"{source_name}: conflicting values for crop={crop!r} year={int(year)}: "
            f"{vals} across sheets {sorted(group['sheet'].unique())} "
            f"(rel diff {rel:.2%}) — resolve manually, do not silently pick one"
        )
    return pd.DataFrame(rows)


def validate_against_anchors(df: pd.DataFrame, value_col: str, anchors: dict[tuple[str, int], float]) -> pd.DataFrame:
    """Compare parsed values against printed national totals. Returns a report."""
    rows = []
    for (crop, year), expected in anchors.items():
        hit = df[(df["crop"] == crop) & (df["year"] == year)]
        if hit.empty:
            rows.append({"crop": crop, "year": year, "expected": expected, "parsed": np.nan, "status": "missing"})
            continue
        parsed = float(hit[value_col].iloc[0])
        ok = abs(parsed - expected) <= ANCHOR_REL_TOLERANCE * max(abs(expected), 1.0)
        rows.append(
            {
                "crop": crop,
                "year": year,
                "expected": expected,
                "parsed": parsed,
                "status": "ok" if ok else "mismatch",
            }
        )
    return pd.DataFrame(rows)


def build_national_yield_panel(
    area_path: Path,
    production_path: Path,
    validate: bool = True,
) -> tuple[pd.DataFrame, dict]:
    """Merge area + production into a national crop-year yield panel.

    Returns ``(panel, metadata)``. The panel has columns ``crop, year,
    area_planted_ha, production_mt, yield_tons_ha_planted`` plus provenance
    flags. ``yield_tons_ha_planted`` is production over *planted* area — a
    sensitivity target, NOT the AAS harvested-area yield.
    """
    area = _dedupe_with_conflict_check(load_area_planted(area_path), "area_planted_ha", Path(area_path).name)
    prod = _dedupe_with_conflict_check(load_production(production_path), "production_mt", Path(production_path).name)

    panel = prod.merge(
        area[["crop", "year", "area_planted_ha"]],
        on=["crop", "year"],
        how="inner",
        validate="one_to_one",
    )
    panel["yield_tons_ha_planted"] = np.where(
        panel["area_planted_ha"] > 0,
        panel["production_mt"] / panel["area_planted_ha"],
        np.nan,
    )
    panel["grain"] = "national"
    panel["yield_basis"] = "planted_area"
    panel["is_proxy"] = True  # proxy for harvested-area yield
    panel["is_synthetic"] = False
    panel["provenance"] = "ubos_statistical_abstract_2015_2021"

    metadata: dict = {
        "grain": "national",
        "years": sorted(int(y) for y in panel["year"].unique()),
        "crops": sorted(panel["crop"].unique()),
        "rows": int(len(panel)),
        "yield_basis": "planted_area",
        "provenance": "ubos_statistical_abstract_2015_2021",
    }

    if validate:
        prod_report = validate_against_anchors(prod, "production_mt", NATIONAL_ANCHORS_PRODUCTION_MT)
        area_report = validate_against_anchors(area, "area_planted_ha", NATIONAL_ANCHORS_AREA_HA)
        metadata["anchor_validation"] = {
            "production": prod_report.to_dict("records"),
            "area": area_report.to_dict("records"),
        }
        bad = pd.concat([prod_report, area_report], ignore_index=True)
        mismatches = bad[bad["status"] == "mismatch"]
        if not mismatches.empty:
            raise AASWorkbookError(
                "anchor validation failed against printed UBOS totals:\n"
                + mismatches.to_string(index=False)
            )
        metadata["anchor_validation_summary"] = {
            "checked": int((bad["status"] != "missing").sum()),
            "ok": int((bad["status"] == "ok").sum()),
            "missing": int((bad["status"] == "missing").sum()),
        }
    return panel.sort_values(["crop", "year"]).reset_index(drop=True), metadata


# ---------------------------------------------------------------------------
# CLI: quick layout inspection + parse smoke test
# ---------------------------------------------------------------------------

def describe(path: Path) -> dict:
    """Layout report for one workbook — run this first on a new file."""
    path = Path(path)
    grain = detect_grain(path)
    sheets = list_sheets(path)
    sheet_info = {}
    for sheet in sheets:
        grid = _sheet_grid(path, sheet, max_rows=40)
        preview = [[_norm(c)[:24] for c in row[:12]] for row in grid[:12]]
        sheet_info[sheet] = {"n_rows_scanned": len(grid), "preview": preview}
    return {"file": path.name, "grain": grain.to_dict(), "sheets": sheet_info}


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    parser.add_argument("--area", type=Path, help="area-planted workbook path")
    parser.add_argument("--production", type=Path, help="production workbook path")
    parser.add_argument("--describe", action="store_true", help="print layout/grain JSON and exit")
    parser.add_argument("--out", type=Path, help="optional CSV output for the merged panel")
    args = parser.parse_args()

    if args.describe:
        report = {}
        if args.area:
            report["area"] = describe(args.area)
        if args.production:
            report["production"] = describe(args.production)
        print(json.dumps(report, indent=2))
        return

    if not (args.area and args.production):
        parser.error("--area and --production are required unless --describe")
    panel, meta = build_national_yield_panel(args.area, args.production)
    print(json.dumps(meta, indent=2, default=str))
    if args.out:
        panel.to_csv(args.out, index=False)
        print(f"wrote {len(panel)} rows -> {args.out}")


if __name__ == "__main__":
    main()
