"""Fixture-based tests for the UBOS 2015-2021 Statistical Abstract parser.

The fixtures replicate the published workbook layout (wide crop-by-year
tables with group total rows, multi-sheet production file with overlapping
years) so the parser can be exercised in CI without the raw files.
"""
from __future__ import annotations

import pytest
from openpyxl import Workbook

from uganda_crop_model.data import aas_2015_2021 as mod

YEARS = [2008, 2015, 2016, 2017, 2018, 2019, 2020, 2021]

AREA = {
    "Plantain Bananas (All types)": {2008: 915817, 2015: 973340, 2016: 970250, 2017: 970308, 2018: 578757, 2019: 589000, 2020: 633000, 2021: 2344680},
    "Millet":   {2008: 249987, 2015: 175173, 2016: 175874, 2017: 180000, 2018: 282607, 2019: 230000, 2020: 170000, 2021: 176949},
    "Maize":    {2008: 1014250, 2015: 1125168, 2016: 1128543, 2017: 1500000, 2018: 2480097, 2019: 1887000, 2020: 1854000, 2021: 1265224},
    "Sorghum":  {2008: 399252, 2015: 373382, 2016: 373083, 2017: 334000, 2018: 491911, 2019: 323000, 2020: 332000, 2021: 298791},
    "Rice":     {2008: 75086, 2015: 95277, 2016: 95515, 2017: 132000, 2018: 199538, 2019: 177000, 2020: 199000, 2021: 220634},
    "S/potatoes": {2008: 440256, 2015: 454478, 2016: 454933, 2017: 476000, 2018: 626406, 2019: 503000, 2020: 430000, 2021: 468000},
    "Irish":    {2008: 32759, 2015: 39339, 2016: 39373, 2017: 56000, 2018: 111124, 2019: 94000, 2020: 73000, 2021: 88000},
    "Cassava":  {2008: 871389, 2015: 852340, 2016: 852809, 2017: 852600, 2018: 1705942, 2019: 658000, 2020: 1413000, 2021: 899053},
    "Beans":    {2008: 617522, 2015: 674964, 2016: 674700, 2017: 1100000, 2018: 1205509, 2019: 867000, 2020: 1017000, 2021: 925028},
    "Gnuts":    {2008: 345232, 2015: 422710, 2016: 422836, 2017: 237000, 2018: 515041, 2019: 215000, 2020: 402000, 2021: 315885},
    "Soya Beans": {2008: 36444, 2015: 46688, 2016: 46782, 2017: 46828, 2018: 189710, 2019: 215000, 2020: 190000, 2021: 198000},
}

PROD_SHEET1 = {
    "Plantain Bananas(All Types)": {2015: 4623367, 2018: 6494057, 2019: 9400000, 2020: 8300000, 2021: 11176675},
    "Millet":   {2015: 236484, 2018: 141982, 2019: 226000, 2020: 170000, 2021: 89773},
    "Maize":    {2015: 2812919, 2018: 3442430, 2019: 2800000, 2020: 2750000, 2021: 6164663},
    "Sorghum":  {2015: 410720, 2018: 268493, 2019: 323000, 2020: 299000, 2021: 222449},
    "Rice":     {2015: 238193, 2018: 199266, 2019: 210000, 2020: 373000, 2021: 727120},
    "S/Potatoes": {2015: 1068000, 2018: 1484163, 2019: 1509000, 2020: 1167000, 2021: 1207617},
    "Irish":    {2015: 220000, 2018: 327332, 2019: 300000, 2020: 250000, 2021: 241230},
    "Cassava":  {2015: 2879000, 2018: 4390231, 2019: 3000000, 2020: 2084000, 2021: 2257790},
    "Beans":    {2015: 929000, 2018: 727652, 2019: 627000, 2020: 786000, 2021: 1414574},
    "Gnuts":    {2015: 243000, 2018: 253279, 2019: 215000, 2020: 183000, 2021: 251877},
    "Soya Beans": {2015: 29000, 2018: 107624, 2019: 117000, 2020: 160000, 2021: 265870},
}
PROD_SHEET2 = {  # overlaps 2018/2019 with identical values -> dedupe, no conflict
    "Plantain Bananas(All Types)": {2017: 4616978, 2018: 6494057, 2019: 9400000},
    "Millet":   {2017: 237000, 2018: 141982, 2019: 226000},
    "Maize":    {2017: 2625000, 2018: 3442430, 2019: 2800000},
    "Sorghum":  {2017: 299000, 2018: 268493, 2019: 323000},
    "Rice":     {2017: 188674, 2018: 199266, 2019: 210000},
    "S/Potatoes": {2017: 1094632, 2018: 1484163, 2019: 1509000},
    "Irish":    {2017: 299338, 2018: 327332, 2019: 300000},
    "Cassava":  {2017: 1858666, 2018: 4390231, 2019: 3000000},
    "Beans":    {2017: 746683, 2018: 727652, 2019: 627000},
    "Gnuts":    {2017: 129046, 2018: 253279, 2019: 215000},
    "Soya Beans": {2017: 28097, 2018: 107624, 2019: 117000},
}


def _write_wide_sheet(ws, years, data, title):
    ws.append([title])
    ws.append([])
    ws.append(["Crop"] + years)
    for crop, vals in data.items():
        ws.append([crop] + [vals.get(y) for y in years])
    ws.append(["Total"] + [sum(d.get(y, 0) for d in data.values()) for y in years])


@pytest.fixture()
def workbooks(tmp_path):
    wb = Workbook()
    wb.remove(wb.active)
    cover = wb.create_sheet("Sheet1")
    cover.append(["Area planted for selected Food Crops, 2015-2021(Ha)"])
    _write_wide_sheet(wb.create_sheet("Sheet2"), YEARS, AREA, "Area planted 2015-2021")
    area_path = tmp_path / "area.xlsx"
    wb.save(area_path)

    wb = Workbook()
    wb.remove(wb.active)
    y1 = sorted({y for d in PROD_SHEET1.values() for y in d})
    y2 = sorted({y for d in PROD_SHEET2.values() for y in d})
    _write_wide_sheet(wb.create_sheet("Sheet1"), y1, PROD_SHEET1, "Production 2015-2021")
    _write_wide_sheet(wb.create_sheet("Sheet2"), y2, PROD_SHEET2, "Production cont.")
    prod_path = tmp_path / "production.xlsx"
    wb.save(prod_path)
    return area_path, prod_path


def test_detect_grain_national(workbooks):
    for path in workbooks:
        report = mod.detect_grain(path)
        assert report.grain == "national"


def test_detect_grain_subregion(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.append(["Sub-region", 2018, 2019, 2020])
    for name in ["Acholi", "Ankole", "Lango", "Teso", "Busoga", "Kigezi"]:
        ws.append([name, 1, 2, 3])
    path = tmp_path / "sub.xlsx"
    wb.save(path)
    assert mod.detect_grain(path).grain == "subregion"


def test_parse_area_drops_group_totals(workbooks):
    area_path, _ = workbooks
    area = mod.load_area_planted(area_path)
    assert "total" not in set(area["raw_label"])
    hit = area[(area["crop"] == "maize") & (area["year"] == 2018)]
    assert hit["area_planted_ha"].iloc[0] == 2_480_097
    assert ("banana", 2008) in set(zip(area["crop"], area["year"]))


def test_parse_production_multisheet_overlap(workbooks):
    _, prod_path = workbooks
    prod = mod.load_production(prod_path)
    assert set(prod["sheet"].unique()) == {"Sheet1", "Sheet2"}
    dup = prod.groupby(["crop", "year"]).size()
    assert (dup > 1).any()  # 2018/2019 present on both sheets


def test_build_panel_anchor_validation(workbooks):
    area_path, prod_path = workbooks
    panel, meta = mod.build_national_yield_panel(area_path, prod_path)
    summary = meta["anchor_validation_summary"]
    assert summary["ok"] == summary["checked"]
    assert summary["missing"] == 0
    assert meta["grain"] == "national"
    assert panel["yield_tons_ha_planted"].notna().all()
    maize18 = panel[(panel["crop"] == "maize") & (panel["year"] == 2018)].iloc[0]
    assert maize18["production_mt"] == 3_442_430
    assert abs(maize18["yield_tons_ha_planted"] - 3_442_430 / 2_480_097) < 1e-6
    # provenance flags must travel with the derived target
    assert set(panel["yield_basis"].unique()) == {"planted_area"}
    assert panel["is_proxy"].all() and not panel["is_synthetic"].any()


def test_cross_sheet_conflict_raises(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Crop", 2015, 2016, 2017])
    ws.append(["Maize", 2812919, 3000000, 2625000])
    ws2 = wb.create_sheet("Sheet2")
    ws2.append(["Crop", 2015, 2016, 2017])
    ws2.append(["Maize", 2812919, 9999999, 2625000])  # 2016 conflicts materially
    path = tmp_path / "conflict.xlsx"
    wb.save(path)
    with pytest.raises(mod.AASWorkbookError, match="conflicting values"):
        mod._dedupe_with_conflict_check(mod.parse_workbook(path, "production_mt"), "production_mt", path.name)


def test_transposed_layout(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Year", "Maize", "Beans", "Millet"])
    for y in [2015, 2016, 2017]:
        ws.append([y, 1000 + y, 2000 + y, 3000 + y])
    path = tmp_path / "long.xlsx"
    wb.save(path)
    df = mod.parse_workbook(path, "production_mt")
    assert set(df["crop"]) == {"maize", "beans", "millet"}
    assert len(df) == 9


def test_unparseable_workbook_raises(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.append(["nothing", "resembling", "a", "table"])
    path = tmp_path / "junk.xlsx"
    wb.save(path)
    with pytest.raises(mod.AASWorkbookError, match="no crop/year table"):
        mod.parse_workbook(path, "value")
